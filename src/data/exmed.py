import mysql.connector
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# --- Your DB credentials ---
db_config = {
    'host': 'localhost',
    'user': 'ommal_oummal',
    'password': 'dMR2id57dviMJJnc',
    'database': 'ommal_medlist',
}

# --- Desired column order (edit if needed) ---
columns = [
    'brand_name','atc','seq','bg','ingredients','code','reg_number','strength','presentation','form',
    'dosage_lndi','presentation_lndi','form_lndi','route_lndi','agent','manufacturer','country',
    'public_price','stratum','subsidy_percentage','pill_price','barcode_gtin','added_date',
    'modified_date','modified_by'
]

EXCEL_PATH = "all_medications.xlsx"

# Step 1: Query the database
conn = mysql.connector.connect(**db_config)
query = f"SELECT * FROM medications"
df = pd.read_sql(query, conn)
conn.close()

# Step 2: Reorder columns (add blanks if missing)
for col in columns:
    if col not in df.columns:
        df[col] = ""
df = df[columns]

# Step 3: Write to Excel
df.to_excel(EXCEL_PATH, index=False)

# Step 4: Formatting: filter row, freeze pane, bold headers, autofit columns
wb = load_workbook(EXCEL_PATH)
ws = wb.active

# Add autofilter
ws.auto_filter.ref = ws.dimensions

# Freeze top row
ws.freeze_panes = "A2"

# Header style
header_font = Font(bold=True)
header_alignment = Alignment(horizontal="center")
for col_idx, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
    for cell in col:
        cell.font = header_font
        cell.alignment = header_alignment
    # Autofit column
    max_length = max(
        len(str(cell.value)) if cell.value else 0
        for cell in ws[get_column_letter(col_idx)]
    )
    ws.column_dimensions[get_column_letter(col_idx)].width = max(max_length + 2, 12)

wb.save(EXCEL_PATH)
print(f"Exported all medications to '{EXCEL_PATH}' with formatting.")
