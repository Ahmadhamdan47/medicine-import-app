import os
from copy import copy
from datetime import datetime

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DRUGS_API_URL = "https://apiv2.medleb.org/drugs/all"
EXCLUDED_FIELDS = {
    "dfSequence",
    "DFSequence",
    "FirstLetter",
    "priceUpdateDate",
    "imagesPath",
    "ImagesPath",
}


def _extract_records(payload):
    """Handle common API response wrappers and return a list of records."""
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("data", "rows", "result", "drugs", "items"):
            value = payload.get(key)
            if isinstance(value, list):
                return value
    return []


def _is_na_or_na_string(value):
    """Treat NaN and common N/A strings as missing."""
    if pd.isna(value):
        return True
    if isinstance(value, str) and value.strip().upper() in {"N/A", "NA", ""}:
        return True
    return False


def _is_truly_empty(value):
    """Treat only NaN/empty as missing; a literal NA string counts as filled."""
    if isinstance(value, str):
        return value.strip() == ""
    return pd.isna(value)


def _is_missing_required(field, value):
    # Dosage stores the literal "NA" as a final value, so only empty counts as missing
    if field == "Dosage":
        return _is_truly_empty(value)
    return _is_na_or_na_string(value)


def _ensure_columns(df, columns):
    for col in columns:
        if col not in df.columns:
            df[col] = None
    return df[columns]


def _style_worksheet(ws):
    header_fill = PatternFill(fill_type="solid", fgColor="007BFF")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    if ws.max_row > 0 and ws.max_column > 0:
        ws.auto_filter.ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
    ws.freeze_panes = "A2"

    for column_cells in ws.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 40)


def export_drugtable_from_api(api_url=DRUGS_API_URL, output_folder="exports"):
    """
    Fetch drugs from API and create a local DrugTable workbook with only the default sheet.
    Excludes: dfSequence, FirstLetter, priceUpdateDate.
    Includes only rows where at least one required field is empty.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    os.makedirs(output_folder, exist_ok=True)
    output_file = os.path.join(output_folder, f"DrugTable_All_Presets_{timestamp}.xlsx")

    print("Fetching drugs from API...")
    print(f"URL: {api_url}")

    try:
        response = requests.get(api_url, timeout=120)
        response.raise_for_status()
        payload = response.json()
    except Exception as e:
        print(f"Error fetching API data: {e}")
        return None

    records = _extract_records(payload)
    if not records:
        print("No records returned from API.")
        return None

    print(f"Fetched {len(records)} records")

    normalized_rows = []
    for record in records:
        if not isinstance(record, dict):
            continue

        cleaned_record = {k: v for k, v in record.items() if k not in EXCLUDED_FIELDS}
        presentations = cleaned_record.get("DrugPresentations") or []
        first_presentation = presentations[0] if presentations and isinstance(presentations[0], dict) else {}

        row = {
            "DrugID": cleaned_record.get("DrugID"),
            "DrugName": cleaned_record.get("DrugName"),
            "DrugNameAR": cleaned_record.get("DrugNameAR"),
            "Seq": cleaned_record.get("Seq"),
            "ProductType": cleaned_record.get("ProductType"),
            "ATC": cleaned_record.get("ATC_Code"),
            "ATCRelatedIngredient": cleaned_record.get("ATCRelatedIngredient"),
            "OtherIngredients": cleaned_record.get("OtherIngredients"),
            "Dosage": cleaned_record.get("Dosage"),
            "DosageNumerator1": None,
            "DosageNumerator1Unit": None,
            "Form": cleaned_record.get("Form"),
            "FormRaw": cleaned_record.get("FormRaw"),
            "FormLNDI": cleaned_record.get("FormLNDI"),
            "RouteLNDI": cleaned_record.get("RouteLNDI"),
            "Route": cleaned_record.get("Route"),
            "RouteRaw": cleaned_record.get("RouteRaw"),
            "Parent": cleaned_record.get("RouteParent"),
            "PresentationLNDI": cleaned_record.get("PresentationLNDI"),
            "PresentationUnitQuantity1": first_presentation.get("UnitQuantity1"),
            "PresentationUnitType1": first_presentation.get("UnitType1"),
            "PresentationUnitQuantity2": first_presentation.get("UnitQuantity2"),
            "PresentationUnitType2": first_presentation.get("UnitType2"),
            "PresentationPackageQuantity1": first_presentation.get("PackageQuantity1"),
            "PresentationPackageType1": first_presentation.get("PackageType1"),
            "PresentationPackageQuantity2": first_presentation.get("PackageQuantity2"),
            "PresentationPackageType2": first_presentation.get("PackageType2"),
            "PresentationPackageQuantity3": first_presentation.get("PackageQuantity3"),
            "PresentationPackageType3": first_presentation.get("PackageType3"),
            "PresentationDescription": first_presentation.get("Description") or cleaned_record.get("Presentation"),
            "Stratum": cleaned_record.get("Stratum"),
            "Amount": cleaned_record.get("Amount"),
            "Agent": cleaned_record.get("Agent"),
            "Manufacturer": cleaned_record.get("Manufacturer"),
            "Country": cleaned_record.get("Country"),
            "RegistrationNumber": cleaned_record.get("RegistrationNumber"),
            "PublicPrice": cleaned_record.get("PublicPrice"),
            "SubsidyPercentage": cleaned_record.get("SubsidyPercentage"),
            "Substitutable": cleaned_record.get("Substitutable"),
            "MoPHCode": cleaned_record.get("MoPHCode"),
            "NotMarketed": cleaned_record.get("NotMarketed"),
            "GTIN": cleaned_record.get("GTIN"),
            "isOTC": cleaned_record.get("isOTC"),
        }
        normalized_rows.append(row)

    default_df = pd.DataFrame(normalized_rows)

    default_cols = [
        "DrugID",
        "DrugName",
        "DrugNameAR",
        "Seq",
        "ProductType",
        "ATC",
        "ATCRelatedIngredient",
        "OtherIngredients",
        "Dosage",
        "DosageNumerator1",
        "DosageNumerator1Unit",
        "Form",
        "FormRaw",
        "FormLNDI",
        "RouteLNDI",
        "Route",
        "RouteRaw",
        "Parent",
        "PresentationLNDI",
        "PresentationUnitQuantity1",
        "PresentationUnitType1",
        "PresentationUnitQuantity2",
        "PresentationUnitType2",
        "PresentationPackageQuantity1",
        "PresentationPackageType1",
        "PresentationPackageQuantity2",
        "PresentationPackageType2",
        "PresentationPackageQuantity3",
        "PresentationPackageType3",
        "PresentationDescription",
        "Stratum",
        "Amount",
        "Agent",
        "Manufacturer",
        "Country",
        "RegistrationNumber",
        "PublicPrice",
        "SubsidyPercentage",
        "Substitutable",
        "MoPHCode",
        "NotMarketed",
        "GTIN",
        "isOTC",
    ]
    default_df = _ensure_columns(default_df, default_cols)

    default_only_df = _ensure_columns(default_df.copy(), default_cols)

    required_fields = [
        "ProductType",
        "ATC",
        "OtherIngredients",
        "Dosage",
        "FormRaw",
        "RouteRaw",
        "PresentationLNDI",
    ]

    mask_any_empty = default_only_df[required_fields].apply(
        lambda col: col.map(lambda v: _is_missing_required(col.name, v))
    ).any(axis=1)
    filtered_df = default_only_df[mask_any_empty]

    output_columns = [
        "DrugName",
        "RegistrationNumber",
        "MoPHCode",
        "ProductType",
        "ATC",
        "OtherIngredients",
        "Dosage",
        "FormRaw",
        "RouteRaw",
        "PresentationLNDI",
    ]
    filtered_df = _ensure_columns(filtered_df, output_columns)

    print(
        f"Filtered rows with at least one empty required field: "
        f"{len(filtered_df)} / {len(default_only_df)}"
    )

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        filtered_df.to_excel(writer, sheet_name="default", index=False)
        ws = writer.sheets["default"]
        _style_worksheet(ws)

    print(f"Successfully exported filtered default sheet workbook: {output_file}")
    return output_file


def convert_drugtable_to_tbf(input_file, output_folder="TBF"):
    """
    Convert DrugTable file to TBF format by filtering ATCRelatedIngredient = N/A.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d")
    output_file = os.path.join(output_folder, f"DrugTable_TBF_Format_{timestamp}.xlsx")

    print("Converting DrugTable to TBF format...")
    print(f"Input: {input_file}")
    print(f"Output: {output_file}")
    print("=" * 60)

    try:
        excel_data = pd.read_excel(input_file, sheet_name=None, engine="openpyxl")
        print(f"Found sheets: {list(excel_data.keys())}\n")
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

    filtered_data = {}
    sheets_with_atc = ["default", "substitutionCheck", "atcCheck"]
    filtered_drug_names = set()

    for sheet_name in sheets_with_atc:
        if sheet_name in excel_data:
            df = excel_data[sheet_name]
            print(f"Processing: {sheet_name}")
            print(f"  Original rows: {len(df)}")

            if "ATCRelatedIngredient" in df.columns:
                mask = df["ATCRelatedIngredient"].apply(_is_na_or_na_string)
                filtered_df = df[mask]
                print(f"  Filtered rows (N/A only): {len(filtered_df)}")
                print(f"  Removed rows: {len(df) - len(filtered_df)}")
                filtered_data[sheet_name] = filtered_df

                if sheet_name == "default" and "DrugName" in filtered_df.columns:
                    filtered_drug_names = set(filtered_df["DrugName"].dropna().astype(str))
                    print(f"  Collected {len(filtered_drug_names)} drug names for presentationCheck")
            else:
                print("  Warning: ATCRelatedIngredient column not found")
                filtered_data[sheet_name] = df
        else:
            print(f"Warning: Sheet '{sheet_name}' not found")

    if "presentationCheck" in excel_data:
        df = excel_data["presentationCheck"]
        print("\nProcessing: presentationCheck")
        print(f"  Original rows: {len(df)}")

        if "DrugName" in df.columns and filtered_drug_names:
            filtered_df = df[df["DrugName"].astype(str).isin(filtered_drug_names)]
            print(f"  Filtered rows (matching drug names): {len(filtered_df)}")
            print(f"  Removed rows: {len(df) - len(filtered_df)}")
            filtered_data["presentationCheck"] = filtered_df
        else:
            if "DrugName" not in df.columns:
                print("  Warning: DrugName column not found")
            if not filtered_drug_names:
                print("  Warning: No filtered drug names available")
            filtered_data["presentationCheck"] = df

    print("\n" + "=" * 60)
    print("Saving TBF formatted file...")

    os.makedirs(output_folder, exist_ok=True)
    orig_wb = load_workbook(input_file)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet_name, df in filtered_data.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            if sheet_name in orig_wb.sheetnames:
                orig_ws = orig_wb[sheet_name]

                for col_letter, col_dim in orig_ws.column_dimensions.items():
                    if col_dim.width:
                        ws.column_dimensions[col_letter].width = col_dim.width

                for col_num in range(1, min(orig_ws.max_column + 1, ws.max_column + 1)):
                    orig_cell = orig_ws.cell(row=1, column=col_num)
                    new_cell = ws.cell(row=1, column=col_num)

                    if orig_cell.font:
                        new_cell.font = copy(orig_cell.font)
                    if orig_cell.fill:
                        new_cell.fill = copy(orig_cell.fill)
                    if orig_cell.border:
                        new_cell.border = copy(orig_cell.border)
                    if orig_cell.alignment:
                        new_cell.alignment = copy(orig_cell.alignment)
                    if orig_cell.number_format:
                        new_cell.number_format = orig_cell.number_format

            if ws.max_row > 0 and ws.max_column > 0:
                ws.auto_filter.ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"

            ws.freeze_panes = "A2"

    print("Successfully created TBF format file")
    print(f"Location: {output_file}")
    print("\nSummary:")
    for sheet_name, df in filtered_data.items():
        print(f"  - {sheet_name}: {len(df)} rows")

    return output_file


if __name__ == "__main__":
    # Local run: fetch from web API and create filtered default sheet export.
    export_drugtable_from_api(DRUGS_API_URL)
